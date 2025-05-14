import os
import sys
import threading
import webbrowser  # Para abrir el navegador automáticamente
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import pyodbc
import logging
from functools import wraps
from openpyxl import Workbook
from io import BytesIO
import datetime
import base64
from pathlib import Path
import json
import re
import getpass  # Para obtener el nombre de usuario actual
import socket   # Para obtener el nombre del equipo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill




# Determina si la aplicación está empaquetada con PyInstaller
if getattr(sys, 'frozen', False):
    # Si está empaquetada, usa la ruta temporal proporcionada por PyInstaller
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
else:
    # Si no está empaquetada, usa las rutas normales
    app = Flask(__name__)

CORS(app, resources={
    r"/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})  # Habilitar CORS



# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)






# Configuración de la conexión a SQL Server
server = 'SPICA'
database = 'Auditoria'
conexion_SQL_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes'

server = 'Regulus'
database = 'Source'
conexion_SQL_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes'

server = 'CX5200271957\\Mordor'
database = 'Auditoria'  # Asegúrate de que este sea el nombre correcto
connection_string = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes'



# Variables globales para almacenar la configuración actual de la conexión
PREDEFINED_SERVERS = {
    'SPICA': {'database': 'Auditoria'},
    'Regulus': {'database': 'Source'},
    'CX5200271957\\Mordor': {'database': 'Auditoria'}
}

current_server = 'SPICA'
current_database = PREDEFINED_SERVERS[current_server]['database']
current_user = None
current_password = None

# Endpoint para actualizar la conexión
@app.route('/api/update_connection', methods=['POST'])
def update_connection():
    global current_server, current_database, current_user, current_password
    try:
        data = request.json
        current_server = data.get('server')
        current_database = data.get('database')
        current_user = data.get('user') or None
        current_password = data.get('password') or None
        
        # Verificar la conexión antes de guardar los cambios
        conn = get_db_connection()
        if not conn:
            raise Exception("No se pudo establecer la conexión con los datos proporcionados")
        conn.close()
        
        return jsonify({"message": "Conexión actualizada correctamente."}), 200
    except Exception as e:
        logger.error(f"Error al actualizar la conexión: {e}")
        # Restaurar valores anteriores en caso de error
        current_server = current_server
        current_database = current_database
        current_user = current_user
        current_password = current_password
        return jsonify({"error": str(e)}), 500

# Función para conectar a la base de datos con los parámetros actuales
def get_db_connection(server_override=None, database_override=None, user_override=None, password_override=None):
    try:
        # Usar los parámetros pasados o los globales
        server_to_use = server_override if server_override else current_server
        database_to_use = database_override if database_override else current_database
        user_to_use = user_override if user_override else current_user
        password_to_use = password_override if password_override else current_password

        # Construir la cadena de conexión dinámicamente
        if user_to_use and password_to_use:
            connection_string = (
                f"DRIVER={{SQL Server}};SERVER={server_to_use};DATABASE={database_to_use};"
                f"UID={user_to_use};PWD={password_to_use}"
            )
        else:
            connection_string = (
                f"DRIVER={{SQL Server}};SERVER={server_to_use};DATABASE={database_to_use};"
                f"Trusted_Connection=yes;TrustServerCertificate=yes"
            )
        # Loggear intento de conexión (sin contraseñas)
        safe_connection_string = connection_string.replace(password_to_use or '', '***')
        logger.info(f"Intentando conectar con: {safe_connection_string}")
        conn = pyodbc.connect(connection_string, timeout=30)
        # Verificar la conexión con una consulta simple
        cursor = conn.cursor()
        cursor.execute("SELECT @@VERSION")
        version = cursor.fetchone()
        logger.info(f"Conexión exitosa a SQL Server: {version[0][:50]}...")
        cursor.close()
        return conn
    except Exception as e:
        logger.error(f"Error al conectar a la base de datos: {str(e)}")
        logger.error(f"Server: {server_to_use}, Database: {database_to_use}")
        return None


# Decorador para manejar la conexión automáticamente
def with_db_connection(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        conn = get_db_connection()
        if not conn:
            return jsonify({"error": "Error al conectar a la base de datos"}), 500
        try:
            result = func(conn, *args, **kwargs)
            conn.close()
            return result
        except Exception as e:
            conn.close()
            logger.error(f"Error en la función decorada: {e}")
            return jsonify({"error": "Error interno del servidor"}), 500
    return wrapper

# =========================================================
# Predefined Queries Management
# =========================================================

def load_predefined_queries():
    """
    Loads predefined queries from the queries.json file.
    Returns a dictionary mapping query names to their SQL query strings.
    """
    try:
        json_path = Path(__file__).parent / 'queries.json'
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"Error loading predefined queries: {e}")
        return {}

def save_predefined_queries(queries_data):
    """
    Saves predefined queries to the queries.json file.
    Returns True on success, False on failure.
    """
    try:
        json_path = Path(__file__).parent / 'queries.json'
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(queries_data, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        logger.error(f"Error saving predefined queries: {e}")
        return False

# Función modificada para obtener información de las columnas editables
def get_query_editable_columns(query_name):
    """
    Gets the list of editable columns for a specific query.
    Returns an empty list if no columns are specifically marked as editable.
    """
    try:
        queries_data = load_predefined_queries()
        if query_name in queries_data:
            return queries_data[query_name].get("editable_columns", [])
        return []
    except Exception as e:
        logger.error(f"Error getting editable columns: {e}")
        return []

# Load predefined queries on startup
queries_data = load_predefined_queries()
PREDEFINED_QUERIES = {}
for name, data in queries_data.items():
    if isinstance(data, dict) and "query" in data:
        PREDEFINED_QUERIES[name] = data["query"]
    else:
        # Fallback for old format or malformed data
        PREDEFINED_QUERIES[name] = data

# =========================================================
# Audit Logging
# =========================================================

def ensure_audit_table_exists(conn):
    """
    Creates the audit table if it doesn't exist.
    Returns True if the table exists or was created successfully, False otherwise.
    """
    try:
        cursor = conn.cursor()
        
        # Check if the table exists
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'AuditoriaTablas' AND SCHEMA_NAME(schema_id) = 'dbo')
            BEGIN
                CREATE TABLE dbo.AuditoriaTablas (
                    ID INT IDENTITY(1,1) PRIMARY KEY,
                    Fecha DATETIME DEFAULT GETDATE(),
                    Usuario VARCHAR(255) NOT NULL,
                    Accion VARCHAR(50) NOT NULL,
                    TablaAfectada VARCHAR(255) NOT NULL,
                    DatosAnteriores NVARCHAR(MAX) NULL,
                    DatosNuevos NVARCHAR(MAX) NULL,
                    IP VARCHAR(50) NULL,
                    Aplicacion VARCHAR(255) DEFAULT 'Consultas Web'
                )
            END
        """)
        conn.commit()
        
        # Verify the table was created correctly
        try:
            cursor.execute("SELECT TOP 1 * FROM dbo.AuditoriaTablas")
            cursor.fetchone()  # Just to validate we can access the table
            logger.info("Audit table verified successfully")
        except Exception as e:
            logger.error(f"Table exists but cannot be accessed: {e}")
            return False
            
        return True
    except Exception as e:
        logger.error(f"Error creating audit table: {e}")
        # Print more detailed error information
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return False

def log_audit_record(conn, action, table, old_data=None, new_data=None):
    """
    Logs a record to the audit table.
    Returns True on success, False on failure.
    """
    try:
        # Ensure the audit table exists
        ensure_audit_table_exists(conn)
        
        cursor = conn.cursor()
        
        # Obtener el usuario que realizó la acción
        user = get_current_windows_user()
        
        # Get client IP if available
        client_ip = request.remote_addr if request else 'localhost'
        
        # Convert data to JSON if not None
        old_data_json = json.dumps(old_data) if old_data else None
        new_data_json = json.dumps(new_data) if new_data else None
        
        # Log información de depuración
        logger.info(f"Registrando auditoría para: Usuario={user}, Acción={action}, Tabla={table}, IP={client_ip}")
        
        # Insert audit record
        cursor.execute("""
            INSERT INTO AuditoriaTablas (Usuario, Accion, TablaAfectada, DatosAnteriores, DatosNuevos, IP)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user, action, table, old_data_json, new_data_json, client_ip))
        
        conn.commit()
        return True
    except Exception as e:
        logger.error(f"Error logging audit record: {e}")
        return False

# =========================================================
# Basic Routes
# =========================================================

@app.route('/')
def index():
    """Renders the main HTML page."""
    return render_template('index.html')

@app.route('/exit')
def exit_app():
    """Route to close the application"""
    func = request.environ.get('werkzeug.server.shutdown')
    if func is None:
        raise RuntimeError('Not running with the Werkzeug server')
    func()
    return jsonify({"status": "shutting down"})

# =========================================================
# Query Management Routes
# =========================================================

@app.route('/api/queries', methods=['GET'])
def get_queries():
    """Gets all predefined queries."""
    try:
        json_path = Path(__file__).parent / 'queries.json'
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        return jsonify({})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/queries', methods=['POST'])
def add_query():
    try:
        data = request.json
        name = data.get('name')
        query = data.get('query')
        description = data.get('description', '')
        editable_columns = data.get('editable_columns', [])
        filterable_columns = data.get('filterable_columns', [])
        multiselect_filter_columns = data.get('multiselect_filter_columns', [])
        server = data.get('server')

        # --- Control de acceso por IP ---
        AUTHORIZED_IPS = [
            "",  # IP autorizada para test
        ]
        client_ip = request.remote_addr
        if client_ip not in AUTHORIZED_IPS:
            return jsonify({"error": "No autorizado para agregar consultas"}), 403

        if not name or not query:
            return jsonify({"error": "Name and query are required"}), 400

        json_path = Path(__file__).parent / 'queries.json'
        current_queries = {}
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                current_queries = json.load(f)

        if name in current_queries:
            return jsonify({"error": "Query name already exists"}), 400

        current_queries[name] = {
            "description": description,
            "query": query,
            "editable_columns": editable_columns,
            "filterable_columns": filterable_columns,
            "multiselect_filter_columns": multiselect_filter_columns,
            "server": server
        }

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(current_queries, f, indent=4, ensure_ascii=False)

        # Update in-memory queries
        global PREDEFINED_QUERIES
        PREDEFINED_QUERIES = load_predefined_queries()

        return jsonify({"message": "Query added successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/queries/<path:name>', methods=['DELETE'])
def delete_query(name):
    """Deletes a predefined query."""
    try:
        # --- Control de acceso por IP ---
        AUTHORIZED_IPS = [
            "",  # IP autorizada para test
        ]
        client_ip = request.remote_addr
        if client_ip not in AUTHORIZED_IPS:
            return jsonify({"error": "No autorizado para borrar consultas"}), 403

        json_path = Path(__file__).parent / 'queries.json'
        with open(json_path, 'r', encoding='utf-8') as f:
            queries = json.load(f)
        
        if name not in queries:
            return jsonify({"error": "Query not found"}), 404
            
        del queries[name]
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(queries, f, indent=4, ensure_ascii=False)
            
        # Update in-memory queries
        global PREDEFINED_QUERIES
        PREDEFINED_QUERIES = load_predefined_queries()
        
        return jsonify({"message": "Query deleted successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/queries/<path:name>', methods=['PUT'])
def update_query(name):
    """Updates an existing predefined query."""
    try:
        data = request.json
        query = data.get('query')
        description = data.get('description')
        editable_columns = data.get('editable_columns', [])
        filterable_columns = data.get('filterable_columns', [])
        multiselect_filter_columns = data.get('multiselect_filter_columns', [])

        if not query:
            return jsonify({"error": "Query is required"}), 400
            
        json_path = Path(__file__).parent / 'queries.json'
        with open(json_path, 'r', encoding='utf-8') as f:
            queries = json.load(f)
            
        if name not in queries:
            return jsonify({"error": "Query not found"}), 404
            
        queries[name] = {
            "description": description or queries[name].get("description", ""),
            "query": query,
            "editable_columns": editable_columns,
            "filterable_columns": filterable_columns,
            "multiselect_filter_columns": multiselect_filter_columns
        }
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(queries, f, indent=4, ensure_ascii=False)
            
        # Update in-memory queries
        global PREDEFINED_QUERIES
        PREDEFINED_QUERIES = load_predefined_queries()
        
        return jsonify({"message": "Query updated successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/get_predefined_queries', methods=['GET'])
def get_predefined_queries():
    """Gets all predefined query names available for the current server."""
    try:
        # Reload queries from file each time
        global PREDEFINED_QUERIES
        PREDEFINED_QUERIES = load_predefined_queries()
        
        # Only show predefined queries for SPICA and Regulus
        if current_server in ['SPICA', 'Regulus']:
            return jsonify(list(PREDEFINED_QUERIES.keys()))
        return jsonify([])  # Return empty list for other servers
    except Exception as e:
        logger.error(f"Error getting predefined queries: {e}")
        return jsonify({"error": str(e)}), 500

def split_cte_and_select(query):
    """
    Separa el bloque CTE (WITH ... ) del SELECT final.
    Devuelve (cte_part, select_part, cte_final_name). Si no es CTE, cte_part será ''.
    Soporta CTEs anidados, comentarios y saltos de línea.
    """
    import re
    q = query.lstrip()
    if not q[:4].upper() == 'WITH':
        return '', query, None
    # Eliminar comentarios de línea -- y de bloque /* */
    q = re.sub(r'--.*?\n', '', q)
    q = re.sub(r'/\*.*?\*/', '', q, flags=re.DOTALL)
    # Buscar el cierre del último paréntesis de los CTEs
    depth = 0
    in_quotes = False
    last_paren = -1
    for i, c in enumerate(q):
        if c == "'":
            in_quotes = not in_quotes
        if not in_quotes:
            if c == '(': depth += 1
            elif c == ')':
                depth -= 1
                if depth == 0:
                    last_paren = i
    if last_paren == -1:
        # No se pudo encontrar el cierre del CTE
        return '', query, None
    # Buscar el primer SELECT después del cierre del último paréntesis
    rest = q[last_paren+1:]
    rest = rest.lstrip(' ;\n\r\t')
    match = re.search(r'(?i)select\b', rest)
    if not match:
        return '', query, None
    select_start = match.start()
    cte_part = q[:last_paren+1]
    select_part = rest[select_start:]
    # Extraer el nombre del CTE final (después de FROM ...)
    cte_final_name = None
    m = re.search(r'from\s+([\w\[\]\."]+)', select_part, re.IGNORECASE)
    if m:
        cte_final_name = m.group(1).strip('[]"')
    return cte_part, select_part, cte_final_name

def clean_select_final(select_part):
    """
    Quita el punto y coma final y separa el ORDER BY del SELECT final.
    Si hay '*' junto con columnas explícitas, deja solo '*'.
    Devuelve (select_clean, order_by_clause)
    """
    select_clean = select_part.strip().rstrip(';')
    order_by = None
    # Buscar ORDER BY fuera de paréntesis (no en subqueries)
    depth = 0
    order_by_pos = -1
    for i in range(len(select_clean)):
        c = select_clean[i]
        if c == '(': depth += 1
        elif c == ')': depth -= 1
        elif depth == 0 and select_clean[i:i+8].upper() == 'ORDER BY':
            order_by_pos = i
            break
    if order_by_pos != -1:
        order_by = select_clean[order_by_pos:].strip()
        select_clean = select_clean[:order_by_pos].strip()
    # Limpiar lista de columnas si hay '*' junto con columnas explícitas
    # Solo si el SELECT es del tipo: SELECT ... FROM ...
    import re
    m = re.match(r'(SELECT\s+)(.+?)(\s+FROM\s+.+)', select_clean, re.IGNORECASE | re.DOTALL)
    if m:
        select_prefix = m.group(1)
        columns_part = m.group(2)
        from_part = m.group(3)
        # Si hay '*' y alguna columna explícita, dejar solo '*'
        columns = [c.strip() for c in columns_part.split(',')]
        if '*' in columns and len(columns) > 1:
            select_clean = f"{select_prefix}*{from_part}"
    return select_clean, order_by

@app.route('/api/execute_predefined_query', methods=['POST'])
def execute_predefined_query():
    """
    Executes a predefined query with filters and pagination.
    Returns the query results, column names, and pagination info.
    """
    conn = None
    try:
        data = request.json
        query_name = data.get('query_name')
        # logger.info(f"Ejecutando consulta: '{query_name}'")  # <-- log para depuración
        page = int(data.get('page', 1))
        page_size = int(data.get('page_size', 100))
        filters = data.get('filters', {})
        order_by = data.get('order_by', None)  # Nuevo parámetro para ordenamiento
        only_columns = data.get('only_columns', False)  # Parámetro para obtener solo las columnas
        
        if not query_name or query_name not in PREDEFINED_QUERIES:
            return jsonify({"error": "Query not found"}), 400
        # Obtener el servidor de la consulta
        queries_data = load_predefined_queries()
        query_info = queries_data.get(query_name, {})
        query_server = query_info.get("server", current_server)
        # Leer el flag de paginación (default: True si no está)
        should_paginate = query_info.get("paginate", True)

        # Usar el servidor de la consulta para la conexión
        conn = get_db_connection(server_override=query_server, database_override=PREDEFINED_SERVERS.get(query_server, {}).get('database', current_database))
        if not conn:
            return jsonify({"error": "Database connection error"}), 500
        cursor = conn.cursor()
        cursor.execute("SET NOCOUNT ON")
        # Get the query text (might be stored as a string or inside a dict)
        if isinstance(PREDEFINED_QUERIES[query_name], dict):
            base_query = PREDEFINED_QUERIES[query_name]["query"]
        else:
            base_query = PREDEFINED_QUERIES[query_name]
        
        # Obtener multiselect_filter_columns
        multiselect_filter_columns = query_info.get("multiselect_filter_columns", [])
        # Obtener filterable_columns
        filterable_columns = query_info.get("filterable_columns", [])
        # Obtener editable_columns
        editable_columns = query_info.get("editable_columns", [])
        
        # Si solo se necesitan las columnas, ejecutar una consulta simplificada
        if only_columns:
            try:
                # Modificar la consulta para obtener solo la estructura (TOP 0)
                base_query_stripped = base_query.lstrip().upper()
                is_cte = base_query_stripped.startswith('WITH')
                
                # Enfoque más seguro para CTEs complejas
                if is_cte:
                    try:
                        # Intentar primero con un enfoque simple para CTEs
                        cte_part, select_part, cte_final_name = split_cte_and_select(base_query)
                        
                        # Usar un enfoque más directo: ejecutar el CTE completo pero limitar a 1 resultado
                        # Esto evita problemas con parámetros y WHERE 1=0
                        simple_query = base_query
                        
                        # Asegurarse de que el SELECT final tenga TOP 1
                        if not re.search(r'TOP\s+\d+', select_part, re.IGNORECASE):
                            select_keyword_match = re.search(r'^SELECT\s+', select_part, re.IGNORECASE)
                            if select_keyword_match:
                                select_keyword_end = select_keyword_match.end()
                                modified_select = select_part[:select_keyword_end] + "TOP 1 " + select_part[select_keyword_end:]
                                simple_query = cte_part + "\n" + modified_select
                        
                        logger.info(f"Obteniendo columnas para CTE con enfoque simplificado")
                        cursor.execute(simple_query)
                        columns = [column[0] for column in cursor.description]
                    except Exception as e_cte:
                        logger.error(f"Error en enfoque simple para CTE: {e_cte}")
                        # Si falla, intentar con el enfoque tradicional
                        test_query = f"SELECT TOP 0 * FROM ({base_query}) AS SubQ"
                        cursor.execute(test_query)
                        columns = [column[0] for column in cursor.description]
                else:
                    # Para consultas no-CTE, usar el enfoque estándar
                    test_query = f"SELECT TOP 0 * FROM ({base_query}) AS SubQ"
                    cursor.execute(test_query)
                    columns = [column[0] for column in cursor.description]
                
                return jsonify({
                    "columns": columns,
                    "filterable_columns": filterable_columns,
                    "editable_columns": editable_columns,
                    "multiselect_filter_columns": multiselect_filter_columns
                })
            except Exception as e:
                logger.error(f"Error obteniendo columnas: {e}")
                # Intentar un enfoque alternativo más simple si el primero falla
                try:
                    logger.info("Intentando método alternativo para obtener columnas")
                    # Ejecutar una consulta simple para obtener la estructura de la tabla
                    if is_cte:
                        # Para CTEs, intentar ejecutar directamente sin modificaciones
                        cursor.execute(base_query)
                        columns = [column[0] for column in cursor.description]
                    else:
                        # Para consultas normales, usar TOP 1
                        cursor.execute(f"SELECT TOP 1 * FROM ({base_query}) AS SubQ")
                        columns = [column[0] for column in cursor.description]
                    
                    return jsonify({
                        "columns": columns,
                        "filterable_columns": filterable_columns,
                        "editable_columns": editable_columns,
                        "multiselect_filter_columns": multiselect_filter_columns
                    })
                except Exception as e2:
                    logger.error(f"Error en método alternativo para obtener columnas: {e2}")
                    return jsonify({"error": f"Error al obtener columnas: {str(e2)}"}), 500
                
                return jsonify({"error": f"Error al obtener columnas: {str(e)}"}), 500
        
        # Process advanced filters
        where_clause, filter_params = build_where_clause(filters, multiselect_filter_columns)
        # Build query with filters
        base_query_stripped = base_query.lstrip().upper()
        is_cte = base_query_stripped.startswith('WITH')
        # Si es CTE, anteponer punto y coma si no lo tiene SOLO si se ejecuta como statement principal
        # Pero si se va a envolver en subquery (para paginación o para obtener columnas), quitar el punto y coma inicial
        if is_cte and base_query.lstrip().startswith(';'):
            base_query = base_query.lstrip()[1:]  # Quita el punto y coma inicial
        if where_clause:
            if is_cte:
                # Usar split_cte_and_select para manejar correctamente las CTEs
                cte_part, select_part, cte_final_name = split_cte_and_select(base_query)
                if cte_final_name:
                    # Enfoque mejorado: Usar el nombre del CTE final para aplicar el filtro
                    # Limpiar el SELECT final para quitar punto y coma y ORDER BY si existe
                    select_clean, order_by_clause = clean_select_final(select_part)
                    
                    # Construir un nuevo SELECT que envuelva el SELECT final con el WHERE
                    if re.search(r'\bFROM\s+', select_clean, re.IGNORECASE):
                        # Si el SELECT ya tiene FROM, envolverlo en un subquery
                        query_with_filters = f"{cte_part}\nSELECT * FROM ({select_clean}) AS FilteredFinalCTE WHERE {where_clause}"
                        
                        # Añadir ORDER BY si existía
                        if order_by_clause:
                            query_with_filters += f" {order_by_clause}"
                    else:
                        # Si por alguna razón no tiene FROM, intentar con el enfoque anterior
                        if re.search(r'\bWHERE\b', select_part, re.IGNORECASE):
                            select_part = re.sub(r'(WHERE[\s\S]+?)(ORDER BY|$)', 
                                                lambda m: m.group(1) + f' AND {where_clause} ' + 
                                                (m.group(2) if m.group(2) else ''), 
                                                select_part, flags=re.IGNORECASE)
                        elif re.search(r'ORDER BY', select_part, re.IGNORECASE):
                            select_part = re.sub(r'(ORDER BY)', f'WHERE {where_clause} \\1', select_part, flags=re.IGNORECASE)
                        else:
                            select_part = select_part.rstrip(';') + f' WHERE {where_clause}'
                        query_with_filters = cte_part + select_part
                else:
                    # Si no se puede determinar el nombre del CTE final, usar el enfoque anterior
                    cte_parts = re.split(r'(SELECT[\s\S]+)$', base_query, flags=re.IGNORECASE)
                    if len(cte_parts) == 3:
                        select_part = cte_parts[2]
                        if re.search(r'\bWHERE\b', select_part, re.IGNORECASE):
                            select_part = re.sub(r'(WHERE[\s\S]+?)(ORDER BY|$)', lambda m: m.group(1) + f' AND {where_clause} ' + (m.group(2) if m.group(2) else ''), select_part, flags=re.IGNORECASE)
                        elif re.search(r'ORDER BY', select_part, re.IGNORECASE):
                            select_part = re.sub(r'(ORDER BY)', f'WHERE {where_clause} \\1', select_part, flags=re.IGNORECASE)
                        else:
                            select_part = select_part.rstrip(';') + f' WHERE {where_clause}'
                        query_with_filters = cte_parts[1] + select_part
                    else:
                        query_with_filters = base_query
            else:
                query_with_filters = f"""
                    SELECT * FROM ({base_query}) AS BaseQuery
                    WHERE {where_clause}
                """
        else:
            query_with_filters = base_query

        # --- LÓGICA DE EJECUCIÓN Y PAGINACIÓN ---
        if not should_paginate:
            # --- Ejecutar SIN paginación --- 
            # logger.info(f"Ejecutando consulta '{query_name}' SIN paginación.")
            try:
                cursor.execute(query_with_filters, filter_params)
                columns = [column[0] for column in cursor.description]
                rows = cursor.fetchall()
                total_records = len(rows)
                current_page = 1
                total_pages = 1
            except Exception as e:
                logger.error(f"Error ejecutando consulta sin paginación: {e}\nSQL: {query_with_filters}")
                return jsonify({"error": f"Error al ejecutar consulta sin paginación: {str(e)}"}), 400
        else:
            # --- Ejecutar CON paginación (lógica existente) --- 
            # logger.info(f"Ejecutando consulta '{query_name}' CON paginación.")
            if is_cte:
      
                cte_part, select_part, cte_final_name = split_cte_and_select(base_query)
                if not select_part.strip().upper().startswith('SELECT'):
                    return jsonify({"error": "No se pudo identificar el SELECT final después del CTE."}), 400
                # Limpiar SELECT final (quitar ';', separar ORDER BY, limpiar SELECT col,*)
                select_clean, order_by_clause = clean_select_final(select_part)
                # 1. OBTENER COLUMNAS:
                #    Construir un query que sea el base_query + SELECT final limpio + WHERE 1=0
                #    para obtener solo la estructura sin errores de sintaxis del driver.
                try:
                    # Usar select_clean (sin ORDER BY y sin ';') y añadir WHERE 1=0
                    select_for_columns = select_clean
                    # Verificar si el select_clean ya tiene un WHERE
                    if re.search(r'\bWHERE\b', select_for_columns, re.IGNORECASE):
                        # Si ya tiene WHERE, añadir AND 1=0
                        # Manejar posible ORDER BY residual si clean_select_final no lo quitó bien (aunque debería)
                        select_for_columns = re.sub(r'(\bORDER\s+BY\s+.*)?$', ' AND 1=0 \1', select_for_columns, flags=re.IGNORECASE | re.DOTALL)
                    else:
                        # Si no tiene WHERE, añadir WHERE 1=0 antes del ORDER BY si existe
                         select_for_columns = re.sub(r'(\bORDER\s+BY\s+.*)?$', ' WHERE 1=0 \1', select_for_columns, flags=re.IGNORECASE | re.DOTALL)
                    
                    # Asegurarse de que no termine con AND 1=0 si no había WHERE original
                    select_for_columns = select_for_columns.replace("WHERE AND 1=0", "WHERE 1=0")

                    query_for_columns_only = f"{cte_part}\n{select_for_columns}"
                    
                    # logger.info(f"Intentando obtener columnas con SQL (WHERE 1=0): {query_for_columns_only}")
                    # IMPORTANTE: No pasar parámetros cuando usamos WHERE 1=0
                    cursor.execute(query_for_columns_only)
                    test_columns = [column[0] for column in cursor.description]
                except Exception as e:
                    logger.error(f"CRÍTICO: No se pudo obtener columnas para paginación de CTE con WHERE 1=0: {e}\nSQL intentado: {query_for_columns_only}")
                    # Intentar el método de fallback con TOP 1 inyectado como último recurso
                    try:
                        # logger.warning("Intentando fallback con TOP 1 inyectado para obtener columnas.")
                        temp_select_part_for_columns = select_part.lstrip()
                        select_keyword_match = re.search(r'^SELECT\s+', temp_select_part_for_columns, re.IGNORECASE)
                        if select_keyword_match:
                            select_keyword_end = select_keyword_match.end()
                            if not re.search(r'^TOP\s+\d+', temp_select_part_for_columns[select_keyword_end:], re.IGNORECASE):
                                modified_select_part_for_columns = temp_select_part_for_columns[:select_keyword_end] + "TOP 1 " + temp_select_part_for_columns[select_keyword_end:]
                            else:
                                modified_select_part_for_columns = temp_select_part_for_columns
                            fallback_query = f"{cte_part}\n{modified_select_part_for_columns}"
                            # logger.info(f"Fallback SQL (TOP 1): {fallback_query}")
                            # IMPORTANTE: No pasar parámetros cuando usamos TOP 1 para obtener columnas
                            cursor.execute(fallback_query)
                            test_columns = [column[0] for column in cursor.description]
                        else:
                            raise ValueError("No se encontró SELECT inicial en fallback.") # Forzar error si tampoco funciona
                    except Exception as e_fallback:
                        logger.error(f"CRÍTICO: Fallback para obtener columnas también falló: {e_fallback}")
                        return jsonify({"error": f"Error crítico al obtener columnas CTE (ambos métodos fallaron): {str(e)}. Verifique sintaxis."}), 400

                # 2. Determinar cláusula ORDER BY para ROW_NUMBER
                final_order_by = order_by_clause # Usar el ORDER BY original del query si existe
                
                # Si hay un order_by en la solicitud, usarlo en lugar del predeterminado
                if order_by and order_by.get('column'):
                    column = order_by.get('column')
                    direction = order_by.get('direction', 'ASC')
                    final_order_by = f"ORDER BY [{column}] {direction}"
                elif not final_order_by:
                    # Si no hay ORDER BY en el query, usar el del JSON si está
                    queries_data = load_predefined_queries()
                    query_info = queries_data.get(query_name, {})
                    json_order_by_col = query_info.get("order_by_column")
                    if json_order_by_col:
                        final_order_by = f"ORDER BY [{json_order_by_col}]"
                    else:
                        # Si no hay nada, usar la primera columna como último recurso
                        fallback_col = next((col for col in test_columns if col and col.upper() != 'ROWNUM'), None)
                        if fallback_col:
                            final_order_by = f"ORDER BY [{fallback_col}]"
                        else:
                             # Si ni siquiera hay columnas, no se puede paginar
                             return jsonify({"error": "La consulta CTE no devuelve columnas para determinar un orden de paginación."}), 400
                # Asegurarse de que final_order_by empiece con ORDER BY
                if not final_order_by.strip().upper().startswith('ORDER BY'):
                     return jsonify({"error": f"Cláusula ORDER BY inválida para paginación: {final_order_by}"}), 400
                # 3. Obtener el total de registros del SELECT final limpio
                count_query = f"{cte_part}\nSELECT COUNT(*) FROM ({select_clean}) AS FinalResult"
                try:
                    # Si hay parámetros de filtro, aplicarlos
                    if where_clause:
                        count_query = f"{cte_part}\nSELECT COUNT(*) FROM ({select_clean}) AS FinalResult WHERE {where_clause}"
                        cursor.execute(count_query, filter_params)
                    else:
                        cursor.execute(count_query)
                    total_records = cursor.fetchone()[0]
                except Exception as e:
                    logger.error(f"Error al contar registros para paginación de CTE: {e}\nSQL: {count_query}")
                    return jsonify({"error": f"Error al contar registros del CTE: {str(e)}"}), 400
                # 4. Calcular los límites de la página
                start_row = (page - 1) * page_size + 1
                end_row = page * page_size
                # 5. Construir el SELECT paginado, usando el ORDER BY determinado
                paginated_select = f'''
                    SELECT * FROM (
                        SELECT *, ROW_NUMBER() OVER ({final_order_by}) as RowNum
                        FROM ({select_clean}) AS FinalSelectResult
                    ) AS PaginatedQuery
                    WHERE RowNum BETWEEN {start_row} AND {end_row}
                '''
                
                # Construir la consulta final con el CTE y el SELECT paginado
                if where_clause:
                    # Si hay filtros, aplicarlos directamente al SELECT final dentro del CTE
                    # Esto asegura que los parámetros se apliquen correctamente
                    filtered_select = f"SELECT * FROM ({select_clean}) AS FilteredResult WHERE {where_clause}"
                    paginated_select = f'''
                        SELECT * FROM (
                            SELECT *, ROW_NUMBER() OVER ({final_order_by}) as RowNum
                            FROM ({filtered_select}) AS FinalSelectResult
                        ) AS PaginatedQuery
                        WHERE RowNum BETWEEN {start_row} AND {end_row}
                    '''
                
                paginated_query = f"{cte_part}\n{paginated_select}"
                # logger.info(f"[CTE PAGINADO FINAL] Usando {final_order_by}\nSQL ejecutado:\n{paginated_query}")
                try:
                    cursor.execute(paginated_query, filter_params)
                    columns = [column[0] for column in cursor.description if column[0] != 'RowNum']
                    rows = cursor.fetchall()
                except Exception as e:
                    logger.error(f"Error ejecutando paginación de CTE: {e}\nSQL: {paginated_query}")
                    return jsonify({"error": f"Error al ejecutar paginación CTE: {str(e)}"}), 400
                current_page = page
                total_pages = -(-total_records // page_size)
            else:
                # PAGINACIÓN PARA queries normales
                # Get total records
                count_query = f"SELECT COUNT(*) FROM ({query_with_filters}) AS CountQuery"
                cursor.execute(count_query, filter_params)
                total_records = cursor.fetchone()[0]

                # Determinar el ORDER BY para la paginación
                order_by_clause = ""
                if order_by and order_by.get('column'):
                    column = order_by.get('column')
                    direction = order_by.get('direction', 'ASC')
                    order_by_clause = f"ORDER BY [{column}] {direction}"
                else:
                    order_by_clause = "ORDER BY (SELECT NULL)"  # Orden predeterminado

                # Apply pagination with custom ORDER BY
                paginated_query = f"""
                    SELECT * FROM (
                        SELECT *, ROW_NUMBER() OVER ({order_by_clause}) as RowNum
                        FROM ({query_with_filters}) AS FilteredQuery
                    ) AS PaginatedQuery
                    WHERE RowNum BETWEEN {(page - 1) * page_size + 1} AND {page * page_size}
                """

                cursor.execute(paginated_query, filter_params)
                columns = [column[0] for column in cursor.description if column[0] != 'RowNum']
                rows = cursor.fetchall()
                current_page = page
                total_pages = -(-total_records // page_size)

        # --- Formatear resultados (común para ambos casos) ---
        editable_columns = get_query_editable_columns(query_name)
        filterable_columns = query_info.get("filterable_columns", [])
        multiselect_filter_columns = query_info.get("multiselect_filter_columns", [])

        data = []
        for row in rows:
            row_dict = {}
            for i, value in enumerate(row[:-1]):  # Exclude RowNum
                if value is None:
                    row_dict[columns[i]] = None
                elif isinstance(value, (datetime.date, datetime.datetime)):
                    row_dict[columns[i]] = value.isoformat()
                else:
                    row_dict[columns[i]] = str(value)
            data.append(row_dict)

        return jsonify({
            "columns": columns,
            "data": data,
            "total_records": total_records,
            "current_page": current_page,
            "total_pages": total_pages,  # Ceiling division
            "isEditable": current_server in ['SPICA', 'CX5200271957\\Mordor'],
            "editable_columns": editable_columns,
            "filterable_columns": filterable_columns,
            "multiselect_filter_columns": multiselect_filter_columns
        })

    except Exception as e:
        logger.error(f"Error executing predefined query: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/download_predefined_query_excel', methods=['POST'])
def download_predefined_query_excel():
    """
    Ejecuta una consulta predefinida y retorna los resultados como archivo Excel.
    Soporta CTEs igual que la paginación.
    """
    conn = None
    try:
        data = request.json
        query_name = data.get('query_name')
        filters = data.get('filters', {})
        order_by = data.get('order_by', None)  # Nuevo parámetro para ordenamiento

        if not query_name or query_name not in PREDEFINED_QUERIES:
            return jsonify({"error": "Query not found"}), 400

        # Obtener el servidor de la consulta
        queries_data = load_predefined_queries()
        query_info = queries_data.get(query_name, {})
        query_server = query_info.get("server", current_server)
        conn = get_db_connection(server_override=query_server, database_override=PREDEFINED_SERVERS.get(query_server, {}).get('database', current_database))
        if not conn:
            return jsonify({"error": "Database connection error"}), 500

        # Get the query text (might be stored as a string or inside a dict)
        if isinstance(PREDEFINED_QUERIES[query_name], dict):
            base_query = PREDEFINED_QUERIES[query_name]["query"]
        else:
            base_query = PREDEFINED_QUERIES[query_name]
        
        cursor = conn.cursor()

        # Process advanced filters
        where_clause, filter_params = build_where_clause(filters)
        
        # Build query with filters
        base_query_stripped = base_query.lstrip().upper()
        is_cte = base_query_stripped.startswith('WITH')
        if is_cte:
            cte_part, select_part, cte_final_name = split_cte_and_select(base_query)
            if not cte_final_name:
                return jsonify({"error": "No se pudo identificar el nombre del CTE final para exportar a Excel."}), 400
                
            # Construir el query base con filtros
            if where_clause:
                # Enfoque mejorado: Usar un subquery para aplicar el filtro al SELECT final
                select_clean, order_by_clause = clean_select_final(select_part)
                
                # Construir un nuevo SELECT que envuelva el SELECT final con el WHERE
                if re.search(r'\bFROM\s+', select_clean, re.IGNORECASE):
                    # Si el SELECT ya tiene FROM, envolverlo en un subquery para aplicar filtros
                    filtered_select = f"SELECT * FROM ({select_clean}) AS FilteredResult WHERE {where_clause}"
                    query = f"{cte_part}\n{filtered_select}"
                else:
                    # Fallback al método anterior
                    query = f"{cte_part}\nSELECT * FROM {cte_final_name} WHERE {where_clause}"
            else:
                query = f"{cte_part}\nSELECT * FROM {cte_final_name}"
                
            # Aplicar ordenamiento si se especificó
            if order_by and order_by.get('column'):
                column = order_by.get('column')
                direction = order_by.get('direction', 'ASC')
                query = f"{query} ORDER BY [{column}] {direction}"
        else:
            # Construir el query base con filtros
            if where_clause:
                query = f"SELECT * FROM ({base_query}) AS BaseQuery WHERE {where_clause}"
            else:
                query = base_query
                
            # Aplicar ordenamiento si se especificó
            if order_by and order_by.get('column'):
                column = order_by.get('column')
                direction = order_by.get('direction', 'ASC')
                query = f"{query} ORDER BY [{column}] {direction}"

        cursor.execute(query, filter_params)
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()

        if not rows:
            return jsonify({"message": "No results found"}), 200

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        # logger.info(f"Asignando nombre de hoja: {sanitize_excel_sheet_name(query_name)} para consulta: {query_name}")
        ws.title = 'Resultados'
        ws.append(columns)
        for row in rows:
            # logger.info(f"Agregando fila al Excel: {row}")
            ws.append(['' if v is None else sanitize_excel_cell_value(v) for v in row])

        # Save to buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{query_name}_{timestamp}.xlsx"

        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/get_table_structure', methods=['POST'])
@with_db_connection
def get_table_structure(conn):
    """
    Gets the structure of the table referenced by a predefined query.
    Returns column names, types, constraints, etc.
    """
    try:
        data = request.json
        query_name = data.get('query_name')
        
        if not query_name or query_name not in PREDEFINED_QUERIES:
            return jsonify({"error": "Query not found"}), 400
            
        # Get the base query
        base_query = PREDEFINED_QUERIES.get(query_name)
        
        # Si es CTE, usar split_cte_and_select para obtener el SELECT final
        base_query_stripped = base_query.lstrip().upper()
        is_cte = base_query_stripped.startswith('WITH')
        if is_cte:
            cte_part, select_part, cte_final_name = split_cte_and_select(base_query)
            test_query = f"{cte_part}\nSELECT TOP 1 * FROM ({select_part}) AS SubQ"
        else:
            test_query = f"SELECT TOP 1 * FROM ({base_query}) AS SubQ"
        # Extract the table name
        table_name = extract_table_name(base_query)
        if not table_name:
            return jsonify({"error": "Could not determine table name"}), 400
            
        # Clean the table name (remove schema, brackets, etc.)
        parts = table_name.split('.')
        schema_name = 'dbo'  # Default schema
        clean_table_name = table_name
        
        # If a schema is specified (format: schema.table)
        if len(parts) > 1:
            schema_name = parts[0].replace('[', '').replace(']', '')
            clean_table_name = parts[1].replace('[', '').replace(']', '')
        else:
            clean_table_name = clean_table_name.replace('[', '').replace(']', '')
        
        # logger.info(f"Looking for table structure: {clean_table_name} in schema: {schema_name}")
        
        cursor = conn.cursor()
        
        # Get table column information
        cursor.execute("""
            SELECT 
                c.COLUMN_NAME, 
                c.DATA_TYPE, 
                c.CHARACTER_MAXIMUM_LENGTH,
                c.IS_NULLABLE,
                CASE WHEN pk.COLUMN_NAME IS NOT NULL THEN 'YES' ELSE 'NO' END AS IS_PRIMARY_KEY,
                COLUMNPROPERTY(OBJECT_ID(c.TABLE_SCHEMA + '.' + c.TABLE_NAME), c.COLUMN_NAME, 'IsIdentity') AS IS_IDENTITY
            FROM 
                INFORMATION_SCHEMA.COLUMNS c
            LEFT JOIN (
                SELECT 
                    ku.TABLE_CATALOG,
                    ku.TABLE_SCHEMA,
                    ku.TABLE_NAME,
                    ku.COLUMN_NAME
                FROM 
                    INFORMATION_SCHEMA.TABLE_CONSTRAINTS AS tc
                JOIN 
                    INFORMATION_SCHEMA.KEY_COLUMN_USAGE AS ku
                    ON tc.CONSTRAINT_TYPE = 'PRIMARY KEY' 
                    AND tc.CONSTRAINT_NAME = ku.CONSTRAINT_NAME
            ) pk 
            ON 
                c.TABLE_CATALOG = pk.TABLE_CATALOG
                AND c.TABLE_SCHEMA = pk.TABLE_SCHEMA
                AND c.TABLE_NAME = pk.TABLE_NAME
                AND c.COLUMN_NAME = pk.COLUMN_NAME
            WHERE 
                c.TABLE_NAME = ? AND c.TABLE_SCHEMA = ?
            ORDER BY 
                c.ORDINAL_POSITION
        """, [clean_table_name, schema_name])
        
        columns_info = []
        for row in cursor.fetchall():
            column_name, data_type, max_length, is_nullable, is_primary_key, is_identity = row
            
            # Format data type with length if applicable
            formatted_type = data_type
            if max_length and max_length > 0 and data_type in ['varchar', 'nvarchar', 'char', 'nchar']:
                formatted_type = f"{data_type}({max_length})"
                
            columns_info.append({
                "name": column_name,
                "type": formatted_type,
                "required": is_nullable == 'NO' and is_identity != 1,  # Not required if identity
                "is_primary_key": is_primary_key == 'YES',
                "is_identity": is_identity == 1
            })
        
        # Get visible columns in the query to mark them
        try:
            cursor.execute(test_query)
            visible_columns = [column[0] for column in cursor.description]
        except Exception as e:
            logger.error(f"Error getting visible columns: {str(e)}")
            visible_columns = []
        
        # Mark which columns are visible in the query
        for column in columns_info:
            column["visible_in_query"] = column["name"] in visible_columns
        
        return jsonify({
            "table_name": f"{schema_name}.{clean_table_name}",
            "columns": columns_info
        })
        
    except Exception as e:
        logger.error(f"Error getting table structure: {str(e)}")
        return jsonify({"error": str(e)}), 500

# Añadir esta función para manejar la obtención del usuario actual
current_client_users = {}  # Diccionario para almacenar los usuarios por IP

@app.route('/api/set_client_user', methods=['POST'])
def set_client_user():
    """
    Establece el usuario del cliente basado en la información enviada desde el navegador
    """
    try:
        data = request.json
        username = data.get('username', '')
        hostname = data.get('hostname', 'Web Client')  # Default value if not provided
        client_ip = request.remote_addr
        
        # Almacenar el usuario con su IP como clave
        current_client_users[client_ip] = f"{username}@{hostname}"
        
        # logger.info(f"Usuario del cliente configurado: {current_client_users[client_ip]} desde IP {client_ip}")
        return jsonify({"message": "Usuario configurado correctamente"}), 200
    except Exception as e:
        logger.error(f"Error al configurar el usuario del cliente: {e}")
        return jsonify({"error": str(e)}), 500

# Modificar la función que obtiene el usuario actual para usar los datos del cliente
def get_current_windows_user():
    try:
        client_ip = request.remote_addr if request else 'localhost'
        
        # Si tenemos información del usuario para esta IP, usarla
        if client_ip in current_client_users:
            return current_client_users[client_ip]
            
        # Método 1: usando getpass.getuser() (usuario del servidor)
        username = getpass.getuser()
        
        # Método 2 (como fallback): usando la variable de entorno USERNAME
        if not username:
            username = os.environ.get('USERNAME')
            
        # Método 3 (como segundo fallback): usando os.getlogin()
        if not username:
            try:
                username = os.getlogin()
            except:
                pass
                
        # Si todo falla, usar un valor por defecto
        if not username:
            username = "Usuario Desconocido"
            
        hostname = socket.gethostname()
        return f"{username}@{hostname} (servidor)"
    except Exception as e:
        logger.error(f"Error al obtener usuario de Windows: {e}")
        return "Usuario Desconocido"

@app.route('/api/update_table', methods=['POST'])
@with_db_connection
def update_table(conn):
    """
    Updates a single field in a database table.
    """
    try:
        data = request.json
        query_name = data.get('query_name')
        updated_row = data.get('updatedRow')
        original_values = data.get('originalValues')

        if not all([query_name, updated_row, original_values]):
            return jsonify({"error": "Required data missing"}), 400

        # Get the base query from queries.json
        # Get the query text (might be stored as a string or inside a dict)
        if isinstance(PREDEFINED_QUERIES.get(query_name), dict):
            base_query = PREDEFINED_QUERIES[query_name]["query"]
        else:
            base_query = PREDEFINED_QUERIES.get(query_name)
        
        if not base_query:
            return jsonify({"error": "Query not found"}), 400

        # Extract the table name from the query
        table_name = extract_table_name(base_query)
        if not table_name:
            return jsonify({"error": "Could not determine table name"}), 400

        cursor = conn.cursor()

        # Get available columns in the query
        cursor.execute(f"SELECT TOP 1 * FROM ({base_query}) AS SubQuery")
        query_columns = [column[0] for column in cursor.description]

        # Get column data types
        cursor.execute("""
            SELECT COLUMN_NAME, DATA_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = ? AND TABLE_SCHEMA = 'dbo'
        """, [table_name.split('.')[-1].replace('[', '').replace(']', '')])
        column_types = {}
        for row in cursor.fetchall():
            column_types[row[0]] = row[1]

        # Identify which field changed
        changed_field = None
        new_value = None
        for key in updated_row.keys():
            if str(updated_row[key]).strip() != str(original_values.get(key, '')).strip():
                changed_field = key
                new_value = updated_row[key]
                break

        if not changed_field:
            return jsonify({"error": "No changes detected"}), 400

        # Get the column type as a string
        column_type = str(column_types.get(changed_field, ''))
        
        # Convert value based on data type
        if column_type.lower() in ['numeric', 'decimal', 'int', 'bigint', 'smallint']:
            try:
                new_value = float(new_value) if new_value.strip() else None
            except ValueError:
                return jsonify({"error": f"Invalid value for field {changed_field}"}), 400
        else:
            new_value = str(new_value).strip()

        # Build WHERE clause using only available columns in the query
        where_conditions = []
        where_params = []
        for col, val in original_values.items():
            if col == changed_field or col not in query_columns:
                continue  # Exclude the updated field and columns not in the query
            
            if val is None or val == '':
                where_conditions.append(f"[{col}] IS NULL")
            else:
                # Get the column type as a string
                col_type = str(column_types.get(col, ''))
                
                # Handle text columns
                if col_type.lower() == 'text':
                    where_conditions.append(f"CAST([{col}] AS NVARCHAR(MAX)) = ?")
                else:
                    where_conditions.append(f"[{col}] = ?")
                
                # Convert WHERE clause value based on type
                if col_type.lower() in ['numeric', 'decimal', 'int', 'bigint', 'smallint']:
                    try:
                        where_params.append(float(val) if val.strip() else None)
                    except (ValueError, AttributeError):
                        where_params.append(val)
                else:
                    where_params.append(str(val).strip())

        # Check that WHERE clause is not empty
        if not where_conditions:
            return jsonify({"error": "Cannot build a valid WHERE clause"}), 400

        # Build and execute UPDATE query
        query = f"""
        UPDATE {table_name} 
        SET [{changed_field}] = ? 
        WHERE {' AND '.join(where_conditions)}
        """
        params = [new_value] + where_params

        # Log query and parameters for debugging
        # logger.info(f"Query: {query}")
        # logger.info(f"Params: {params}")

        cursor.execute(query, params)
        rows_affected = cursor.rowcount
        conn.commit()

        if rows_affected == 0:
            return jsonify({"error": "Record to update not found"}), 404

        # Log the change in the audit table
        change_data = {
            "old_value": original_values.get(changed_field),
            "new_value": new_value,
            "field": changed_field
        }
        log_audit_record(conn, "UPDATE", table_name, original_values, updated_row)

        return jsonify({
            "message": "Record updated successfully",
            "rows_affected": rows_affected
        })

    except Exception as e:
        logger.error(f"Error updating record: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/add_entry', methods=['POST'])
@with_db_connection
def add_entry(conn):
    """
    Adds a new record to a database table.
    """
    try:
        data = request.json
        query_name = data.get('query_name')
        new_row = data.get('newRow')
        
        if not all([query_name, new_row]):
            return jsonify({"error": "Required data missing"}), 400

        # Get the base query from queries.json
        base_query = PREDEFINED_QUERIES.get(query_name)
        if not base_query:
            return jsonify({"error": "Query not found"}), 400

        # Extract the table name from the query
        table_name = extract_table_name(base_query)
        if not table_name:
            return jsonify({"error": "Could not determine table name"}), 400

        cursor = conn.cursor()

        # Get column data types and NULL constraints
        clean_table_name = table_name.split('.')[-1].replace('[', '').replace(']', '')
        cursor.execute("""
            SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = ? AND TABLE_SCHEMA = 'dbo'
        """, [clean_table_name])
        
        column_info = {row[0]: {"type": row[1], "nullable": row[2]} for row in cursor.fetchall()}
        
        # Filter only columns that exist in the table
        filtered_row = {k: v for k, v in new_row.items() if k in column_info}
        
        if not filtered_row:
            return jsonify({"error": "No valid columns provided"}), 400
        
        # Prepare values with correct data types
        processed_values = []
        for column, value in filtered_row.items():
            # Handle NULL values
            if value is None or value.strip() == '':
                if column_info[column]["nullable"] == 'NO':
                    return jsonify({"error": f"Column {column} cannot be NULL"}), 400
                processed_values.append(None)
                continue
                
            # Convert based on data type
            data_type = column_info[column]["type"].lower()
            try:
                if data_type in ['int', 'bigint', 'smallint', 'tinyint']:
                    processed_values.append(int(value))
                elif data_type in ['decimal', 'numeric', 'float', 'real']:
                    processed_values.append(float(value))
                elif data_type in ['bit']:
                    processed_values.append(1 if value.lower() in ['1', 'true', 'yes', 'y'] else 0)
                elif data_type in ['date', 'datetime', 'datetime2', 'smalldatetime']:
                    # Dates will be passed as strings and SQL Server will interpret them
                    processed_values.append(value)
                else:
                    processed_values.append(value)
            except ValueError:
                return jsonify({"error": f"Invalid value for column {column} of type {data_type}"}), 400

        # Build INSERT query
        columns = ", ".join([f"[{column}]" for column in filtered_row.keys()])
        values = ", ".join(["?" for _ in filtered_row])
        query = f"INSERT INTO {table_name} ({columns}) VALUES ({values})"

        # Execute query
        # logger.info(f"Query: {query}")
        # logger.info(f"Values: {processed_values}")
        cursor.execute(query, processed_values)
        conn.commit()

        # Log the new record in the audit table
        log_audit_record(conn, "INSERT", table_name, None, filtered_row)

        return jsonify({"message": "Record added successfully"})
    except Exception as e:
        logger.error(f"Error adding record: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete_entry', methods=['POST'])
@with_db_connection
def delete_entry(conn):
    """
    Deletes a record from a database table.
    """
    try:
        data = request.json
        query_name = data.get('query_name')
        row_values = data.get('rowValues')
        
        if not all([query_name, row_values]):
            return jsonify({"error": "Required data missing"}), 400

        # Get the base query from queries.json
        # Get the query text (might be stored as a string or inside a dict)
        if isinstance(PREDEFINED_QUERIES.get(query_name), dict):
            base_query = PREDEFINED_QUERIES[query_name]["query"]
        else:
            base_query = PREDEFINED_QUERIES.get(query_name)

        if not base_query:
            return jsonify({"error": "Query not found"}), 400

        # Extract the table name from the query
        table_name = extract_table_name(base_query)
        if not table_name:
            return jsonify({"error": "Could not determine table name"}), 400

        cursor = conn.cursor()

        # Get available columns in the query
        cursor.execute(f"SELECT TOP 1 * FROM ({base_query}) AS SubQuery")
        query_columns = [column[0] for column in cursor.description]

        # Get column data types
        cursor.execute("""
            SELECT COLUMN_NAME, DATA_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = ? AND TABLE_SCHEMA = 'dbo'
        """, [table_name.split('.')[-1].replace('[', '').replace(']', '')])
        column_types = {}
        for row in cursor.fetchall():
            column_types[row[0]] = row[1]

        # Build WHERE clause using only available columns in the query
        where_conditions = []
        where_params = []
        for col, val in row_values.items():
            if col not in query_columns:
                continue  # Exclude columns not in the query
            
            if val is None or val == '':
                where_conditions.append(f"[{col}] IS NULL")
            else:
                # Get the column type as a string
                col_type = str(column_types.get(col, ''))
                
                # Handle text columns
                if col_type.lower() == 'text':
                    where_conditions.append(f"CAST([{col}] AS NVARCHAR(MAX)) = ?")
                else:
                    where_conditions.append(f"[{col}] = ?")
                
                # Convert WHERE clause value based on type
                if col_type.lower() in ['numeric', 'decimal', 'int', 'bigint', 'smallint']:
                    try:
                        where_params.append(float(val) if val.strip() else None)
                    except (ValueError, AttributeError):
                        where_params.append(val)
                else:
                    where_params.append(str(val).strip())

        # Check that WHERE clause is not empty
        if not where_conditions:
            return jsonify({"error": "Cannot build a valid WHERE clause"}), 400

        # Build and execute DELETE query
        query = f"""
        DELETE FROM {table_name} 
        WHERE {' AND '.join(where_conditions)}
        """

        # Log query and parameters for debugging
        # logger.info(f"Delete query: {query}")
        # logger.info(f"Params: {where_params}")

        cursor.execute(query, where_params)
        rows_affected = cursor.rowcount
        conn.commit()

        if rows_affected == 0:
            return jsonify({"error": "Record to delete not found"}), 404

        # Log the deletion in the audit table
        log_audit_record(conn, "DELETE", table_name, row_values, None)

        return jsonify({
            "message": "Record deleted successfully",
            "rows_affected": rows_affected
        })

    except Exception as e:
        logger.error(f"Error deleting record: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/save_excel', methods=['POST'])
def save_excel():
    """
    Saves an Excel file from base64-encoded data to the user's Downloads folder.
    """
    try:
        data = request.json
        if not data or 'data' not in data or 'filename' not in data:
            return jsonify({"error": "Incomplete data"}), 400

        # Decode base64
        excel_data = base64.b64decode(data['data'])
        
        # Define downloads folder
        downloads_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
        if not os.path.exists(downloads_folder):
            downloads_folder = os.path.dirname(os.path.abspath(__file__))

        # Create full file path
        file_path = os.path.join(downloads_folder, data['filename'])
        
        # Save the file
        with open(file_path, 'wb') as f:
            f.write(excel_data)

        return jsonify({"message": "File saved successfully", "path": file_path}), 200

    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        return jsonify({"error": str(e)}), 500

# =========================================================
# Audit Records Routes
# =========================================================

@app.route('/api/audit/records', methods=['POST'])
@with_db_connection
def get_audit_records(conn):
    """
    Gets audit records with filters and pagination.
    This data is read-only and cannot be modified via the UI.
    """
    try:
        data = request.json
        page = int(data.get('page', 1))
        page_size = int(data.get('page_size', 10))
        filters = data.get('filters', {})
        
        # Build query with filters
        query = """
            SELECT 
                ID, 
                Fecha, 
                Usuario, 
                Accion, 
                TablaAfectada, 
                DatosAnteriores, 
                DatosNuevos, 
                IP, 
                Aplicacion 
            FROM 
                [Auditoria].[dbo].[AuditoriaTablas]
            WHERE 
                1=1
        """
        
        params = []
        
        # Apply filters
        if filters.get('usuario'):
            query += " AND Usuario LIKE ?"
            params.append(f"%{filters['usuario']}%")
            
        if filters.get('accion'):
            query += " AND Accion = ?"
            params.append(filters['accion'])
            
        if filters.get('tabla'):
            query += " AND TablaAfectada LIKE ?"
            params.append(f"%{filters['tabla']}%")
            
        if filters.get('fecha'):
            # Format date as YYYY-MM-DD for SQL Server
            query += " AND CONVERT(date, Fecha) = ?"
            params.append(filters['fecha'])
        
        # Count total records
        count_query = f"""
            SELECT COUNT(*) FROM (
                {query}
            ) AS CountQuery
        """
        
        cursor = conn.cursor()
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()[0]
        
        # Add order by and pagination
        query += " ORDER BY Fecha DESC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
        offset = (page - 1) * page_size
        params.append(offset)
        params.append(page_size)
        
        # Execute query
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        
        # Format results
        result = []
        for row in rows:
            row_dict = {}
            for i, value in enumerate(row):
                if value is None:
                    row_dict[columns[i]] = None
                elif isinstance(value, (datetime.date, datetime.datetime)):
                    row_dict[columns[i]] = value.isoformat()
                elif columns[i] in ['DatosAnteriores', 'DatosNuevos'] and value:
                    # For JSON data stored as strings, keep them as is
                    row_dict[columns[i]] = value
                else:
                    row_dict[columns[i]] = str(value)
            result.append(row_dict)
        
        return jsonify({
            "records": result,
            "total": total_records,
            "page": page,
            "page_size": page_size,
            "total_pages": -(-total_records // page_size)  # Ceiling division
        })
        
    except Exception as e:
        logger.error(f"Error fetching audit records: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/audit/excel', methods=['POST'])
@with_db_connection
def download_audit_excel(conn):
    """
    Exports audit records to Excel based on filters.
    """
    try:
        data = request.json
        filters = data.get('filters', {})
        
        # Build query with filters
        query = """
            SELECT 
                ID, 
                Fecha, 
                Usuario, 
                Accion, 
                TablaAfectada, 
                DatosAnteriores, 
                DatosNuevos, 
                IP, 
                Aplicacion 
            FROM 
                [Auditoria].[dbo].[AuditoriaTablas]
            WHERE 
                1=1
        """
        
        params = []
        
        # Apply filters
        if filters.get('usuario'):
            query += " AND Usuario LIKE ?"
            params.append(f"%{filters['usuario']}%")
            
        if filters.get('accion'):
            query += " AND Accion = ?"
            params.append(filters['accion'])
            
        if filters.get('tabla'):
            query += " AND TablaAfectada LIKE ?"
            params.append(f"%{filters['tabla']}%")
            
        if filters.get('fecha'):
            # Format date as YYYY-MM-DD for SQL Server
            query += " AND CONVERT(date, Fecha) = ?"
            params.append(filters['fecha'])
        
        # Add order by
        query += " ORDER BY Fecha DESC"
        
        # Execute query
        cursor = conn.cursor()
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        
        if not rows:
            return jsonify({"message": "No audit records found"}), 200
            
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros de Auditoría"
        
        # Add header row with formatting
        header_fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
        header_font = Font(bold=True)
        
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Add data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                if value is None:
                    ws.cell(row=row_idx, column=col_idx, value="")
                elif isinstance(value, (datetime.date, datetime.datetime)):
                    ws.cell(row=row_idx, column=col_idx, value=value.isoformat())
                else:
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # Save to buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Registros_Auditoria_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error generating audit Excel: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/unique_filter_values', methods=['POST'])
def unique_filter_values():
    """
    Devuelve los valores únicos de un campo filtrable para una consulta predefinida.
    """
    try:
        data = request.json
        query_name = data.get('query_name')
        column = data.get('column')
        if not query_name or not column:
            return jsonify({"error": "Faltan parámetros"}), 400

        queries_data = load_predefined_queries()
        query_info = queries_data.get(query_name)
        if not query_info:
            return jsonify({"error": "Consulta no encontrada"}), 404

        multiselect_filter_columns = query_info.get('multiselect_filter_columns', [])
        if column not in multiselect_filter_columns:
            return jsonify({"error": "Columna no permitida para multi-select"}), 400

        # Obtener el SQL base y el servidor
        base_query = query_info['query']
        query_server = query_info.get("server", current_server)
        
        # Verificar si es una consulta CTE
        base_query_stripped = base_query.lstrip().upper()
        is_cte = base_query_stripped.startswith('WITH')
        
        # Construir la consulta para valores únicos
        if is_cte:
            # Para consultas CTE, usar split_cte_and_select para manejar correctamente la estructura
            cte_part, select_part, cte_final_name = split_cte_and_select(base_query)
            
            if cte_final_name:
                # Si se pudo extraer el nombre del CTE final, usarlo directamente
                sql = f"{cte_part}\nSELECT DISTINCT [{column}] FROM {cte_final_name} WHERE [{column}] IS NOT NULL ORDER BY [{column}]"
            else:
                # Si no se pudo extraer el nombre, usar el SELECT final en un subquery
                select_clean, _ = clean_select_final(select_part)
                sql = f"{cte_part}\nSELECT DISTINCT [{column}] FROM ({select_clean}) AS MultiSelectSubQ WHERE [{column}] IS NOT NULL ORDER BY [{column}]"
        else:
            # Para consultas normales, usar un subquery
            sql = f"SELECT DISTINCT [{column}] FROM ({base_query}) AS SubQ WHERE [{column}] IS NOT NULL ORDER BY [{column}]"

        # Conectar a la base de datos usando el servidor específico de la consulta
        conn = get_db_connection(server_override=query_server, database_override=PREDEFINED_SERVERS.get(query_server, {}).get('database', current_database))
        if not conn:
            return jsonify({"error": "Error de conexión a la base de datos"}), 500
            
        cursor = conn.cursor()
        cursor.execute(sql)
        values = [row[0] for row in cursor.fetchall() if row[0] is not None]
        cursor.close()
        conn.close()
        return jsonify({"values": values})
    except Exception as e:
        logger.error(f"Error en unique_filter_values: {e}")
        return jsonify({"error": str(e)}), 500

# =========================================================
# Utility Functions
# =========================================================

def open_browser():
    """Opens the default web browser to the application URL."""
    url = "http://127.0.0.1:3000"
    threading.Timer(1, lambda: webbrowser.open(url)).start()

def is_server_editable(server):
    """Determines if a server allows editing data."""
    return server in ['SPICA', 'CX5200271957\\Mordor']

def extract_table_name(query):
    """
    Extracts the table name from a SQL query.
    Enhanced to handle complex cases.
    """
    query = query.lower()

    # Common patterns for extracting table names
    patterns = [
        # Basic pattern: FROM table
        r"from\s+([^\s,()]+)",
        # Pattern for: FROM [schema].[table]
        r"from\s+\[?(\w+)\]?\.\[?(\w+)\]?",
        # Pattern for: FROM table AS alias
        r"from\s+([^\s]+)\s+as\s+\w+",
        # Pattern for join
        r"join\s+([^\s,()]+)"
    ]

    # Try each pattern
    for pattern in patterns:
        matches = re.search(pattern, query)
        if matches:
            if len(matches.groups()) == 1:
                # Only table
                return matches.group(1)
            elif len(matches.groups()) == 2:
                # Schema and table
                return f"{matches.group(1)}.{matches.group(2)}"

    # Fallback method
    if "from" in query:
        after_from = query.split("from")[1].strip()
        table_parts = after_from.split()[0].split('.')
        
        # Handle schema.table case
        if len(table_parts) > 1:
            return '.'.join(table_parts[:2])  # Take only schema.table
        else:
            return table_parts[0]
            
    return None

def build_where_clause(filters, multiselect_filter_columns=None):
    """
    Builds a SQL WHERE clause and parameter list from a dict of filters.
    Handles various filter formats including comparison operators, IN, NULL, etc.
    """
    if not filters:
        return "", []
    if multiselect_filter_columns is None:
        multiselect_filter_columns = []
    conditions = []
    params = []
    for column, filter_value in filters.items():
        if not filter_value.strip():
            continue
        or_conditions = filter_value.split(' OR ')
        or_parts = []
        for condition in or_conditions:
            condition = condition.strip()
            # IN automático: si hay comas y no es un operador ni IN(...)
            if ',' in condition and not (
                condition.startswith('=') or
                condition.startswith('>=') or
                condition.startswith('<=') or
                condition.startswith('>') or
                condition.startswith('<') or
                condition.upper().startswith('LIKE') or
                condition.upper().startswith('IN(')
            ):
                values = [v.strip() for v in condition.split(',') if v.strip()]
                if len(values) > 1:
                    placeholders = ','.join(['?' for _ in values])
                    or_parts.append(f"{column} IN ({placeholders})")
                    params.extend(values)
                    continue
            # Igualdad exacta
            if condition.startswith('='):
                value = condition[1:].strip()
                or_parts.append(f"{column} = ?")
                params.append(value)
                continue
            # Mayor o igual
            if condition.startswith('>='):
                value = condition[2:].strip()
                or_parts.append(f"{column} >= ?")
                params.append(value)
                continue
            # Menor o igual
            if condition.startswith('<='):
                value = condition[2:].strip()
                or_parts.append(f"{column} <= ?")
                params.append(value)
                continue
            # Mayor que
            if condition.startswith('>'):
                value = condition[1:].strip()
                or_parts.append(f"{column} > ?")
                params.append(value)
                continue
            # Menor que
            if condition.startswith('<'):
                value = condition[1:].strip()
                or_parts.append(f"{column} < ?")
                params.append(value)
                continue
            # LIKE explícito
            if condition.upper().startswith('LIKE'):
                value = condition[4:].strip()
                or_parts.append(f"{column} LIKE ?")
                params.append(value)
                continue
            # IN explícito
            if condition.upper().startswith('IN(') and condition.endswith(')'):
                values = [v.strip() for v in condition[3:-1].split(',') if v.strip()]
                if values:
                    placeholders = ','.join(['?' for _ in values])
                    or_parts.append(f"{column} IN ({placeholders})")
                    params.extend(values)
                continue
            # NULL explícito
            if condition.upper() == 'NULL' or condition.upper() == 'IS NULL':
                or_parts.append(f"{column} IS NULL")
                continue
            # Si llegamos aquí, usar LIKE como fallback
            or_parts.append(f"{column} LIKE ?")
            params.append(f"%{condition}%")
        if or_parts:
            conditions.append('(' + ' OR '.join(or_parts) + ')')
    where_clause = ' AND '.join(conditions)
    # logger.info(f"[FINAL WHERE] {where_clause}")
    # logger.info(f"[FINAL PARAMS] {params}")
    return where_clause, params

def sanitize_excel_sheet_name(name):
    # Quita caracteres no permitidos por Excel
    name = re.sub(r'[:\\/?*\[\]\u0000-\u001F]', '', name)
    # Quita caracteres de control y recorta a 31 caracteres
    name = ''.join(c for c in name if 32 <= ord(c) <= 126 or c.isalnum() or c in ' _-').strip()
    if not name:
        name = 'Hoja1'
    return name[:31]

def sanitize_excel_cell_value(value):
    if isinstance(value, str):
        # Elimina caracteres de control excepto salto de línea (\n) y tabulación (\t)
        return ''.join(c for c in value if (ord(c) >= 32 or c in '\n\t'))
    return value

# =========================================================
# Server Startup
# =========================================================

if __name__ == '__main__':
    # Production configuration
    from waitress import serve
    
    # Get server IP
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    
    # Configure detailed logging
    logging.basicConfig(
        filename='app.log',
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Verify initial connection
    logger.info("Verifying initial SQL Server connection...")
    conn = get_db_connection()
    if conn:
        logger.info("Initial connection successful")
        conn.close()
    else:
        logger.error("Could not establish initial connection")
    
    # Show access information
    print(f"""
    Server started:
    - Local URL: http://{local_ip}:3000
    - Hostname URL: http://{hostname}:3000
    
    The application is available on the local network.
    Press Ctrl+C to stop the server.
    """)
    
    # Start server with Waitress on port 3000
    serve(app, host='0.0.0.0', port=3000, threads=8)